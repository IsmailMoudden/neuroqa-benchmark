"""
Main benchmark runner.
Runs 4 chunking strategies × 16 questions and saves results.
Uses OpenRouter (OpenAI-compatible) for LLM calls.
"""

import os
import json
import time
import random
from pathlib import Path
from typing import List, Dict, Any
from concurrent.futures import ThreadPoolExecutor, as_completed

import urllib.request
import numpy as np
from sentence_transformers import SentenceTransformer
import openpyxl
from openpyxl.styles import PatternFill, Font

from chunking_strategies import (
    load_docx_text,
    load_docx_structured,
    sliding_window_chunks,
    structure_aware_chunks,
    semantic_chunks,
)
from metrics import (
    is_relevant_chunk,
    recall_at_k,
    mrr,
    token_f1,
    faithfulness_score,
    relevance_score,
    compute_cost,
)
from qa_dataset import QA_DATASET


# ─── .env loader ──────────────────────────────────────────────────────────────

def _load_dotenv():
    env_file = Path(__file__).parent / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

_load_dotenv()

# ─── Configuration ────────────────────────────────────────────────────────────

RANDOM_SEED = 42
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

DOCS_DIR = Path(__file__).parent / "sources "
RESULTS_DIR = Path(__file__).parent / "results"
RESULTS_DIR.mkdir(exist_ok=True)

EMBED_MODEL_NAME = "all-MiniLM-L6-v2"
_EMBED_MODEL_CACHE = None
LLM_MODEL = "google/gemma-3-4b-it:free"
TOP_K = 5

STRATEGIES = [
    {
        "id": "C",
        "name": "sliding_window_512_128",
        "type": "sliding_window",
        "params": {"chunk_size": 512, "overlap": 128},
    },
    {
        "id": "D",
        "name": "sliding_window_512_256",
        "type": "sliding_window",
        "params": {"chunk_size": 512, "overlap": 256},
    },
    {
        "id": "E",
        "name": "structure_aware_512",
        "type": "structure_aware",
        "params": {"max_tokens": 512, "min_tokens": 50},
    },
    {
        "id": "F",
        "name": "semantic_075_512",
        "type": "semantic",
        "params": {"threshold": 0.75, "max_tokens": 512, "embed_model": EMBED_MODEL_NAME},
    },
]

ANSWER_PROMPT = (
    "You are a helpful assistant. Using ONLY the context below, answer the question concisely.\n\n"
    "Context:\n{context}\n\nQuestion: {question}\n\nAnswer:"
)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def load_all_docs() -> Dict[str, str]:
    docs = {}
    for f in DOCS_DIR.iterdir():
        if f.suffix.lower() == ".docx":
            docs[f.name] = load_docx_text(str(f))
    print(f"Loaded {len(docs)} documents: {list(docs.keys())}")
    return docs


def build_chunks(strategy: Dict, docs: Dict[str, str]) -> List[Dict]:
    all_chunks = []
    stype = strategy["type"]
    params = strategy["params"]

    for doc_name, doc_text in docs.items():
        if stype == "sliding_window":
            chunks = sliding_window_chunks(
                doc_text,
                chunk_size=params["chunk_size"],
                overlap=params["overlap"],
                doc_name=doc_name,
            )
        elif stype == "structure_aware":
            elements = load_docx_structured(str(DOCS_DIR / doc_name))
            chunks = structure_aware_chunks(
                elements,
                max_tokens=params["max_tokens"],
                min_tokens=params["min_tokens"],
                doc_name=doc_name,
            )
        elif stype == "semantic":
            chunks = semantic_chunks(
                doc_text,
                threshold=params["threshold"],
                max_tokens=params["max_tokens"],
                embed_model_name=params["embed_model"],
                doc_name=doc_name,
            )
        else:
            raise ValueError(f"Unknown strategy type: {stype}")
        all_chunks.extend(chunks)

    return all_chunks


def embed_chunks(chunks: List[Dict], model: SentenceTransformer) -> np.ndarray:
    texts = [c["text"] for c in chunks]
    return model.encode(texts, show_progress_bar=False, normalize_embeddings=True, batch_size=64)


def retrieve_top_k(
    query: str,
    chunks: List[Dict],
    chunk_embeddings: np.ndarray,
    model: SentenceTransformer,
    k: int = 5,
) -> List[Dict]:
    q_emb = model.encode([query], normalize_embeddings=True)[0]
    scores = chunk_embeddings @ q_emb
    top_indices = np.argsort(scores)[::-1][:k]
    return [chunks[i] for i in top_indices]


def _openrouter_chat(api_key: str, messages: list, max_tokens: int = 512) -> dict:
    """Direct HTTP call to OpenRouter — bypasses openai SDK auth issues."""
    payload = json.dumps({
        "model": LLM_MODEL,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": 0,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://neuroqa-benchmark.streamlit.app",
            "X-Title": "NeuroQA Benchmark",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code}: {body}") from e


def generate_answer(
    api_key: str,
    question: str,
    context_chunks: List[Dict],
) -> tuple[str, int, int]:
    context = "\n\n---\n\n".join(c["text"] for c in context_chunks)
    prompt = ANSWER_PROMPT.format(context=context, question=question)
    messages = [{"role": "user", "content": prompt}]

    for attempt in range(3):
        try:
            data = _openrouter_chat(api_key, messages, max_tokens=512)
            answer = data["choices"][0]["message"]["content"].strip()
            usage = data.get("usage", {})
            tokens_in = usage.get("prompt_tokens", 0)
            tokens_out = usage.get("completion_tokens", 0)
            return answer, tokens_in, tokens_out
        except Exception as e:
            if attempt < 2:
                print(f"  [generate] retry {attempt+1}: {type(e).__name__}: {e}")
                time.sleep(2 ** attempt)
            else:
                print(f"  [generate] FAILED after 3 attempts: {type(e).__name__}: {e}")
                raise


# ─── Main ─────────────────────────────────────────────────────────────────────

def run_benchmark(strategies: List[Dict] = None, embed_model: SentenceTransformer = None, api_key: str = None, questions: List[Dict] = None):
    if strategies is None:
        strategies = STRATEGIES
    if api_key is None:
        api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        raise ValueError("OPENROUTER_API_KEY not set. Add it to .env")
    if questions is None:
        questions = QA_DATASET
    if not questions:
        raise ValueError("No questions to evaluate. Add questions in the Questions page.")


    # Use passed-in model (cached by caller) or load fresh
    if embed_model is None:
        global _EMBED_MODEL_CACHE
        if _EMBED_MODEL_CACHE is None:
            print(f"Loading embedding model {EMBED_MODEL_NAME}...")
            _EMBED_MODEL_CACHE = SentenceTransformer(EMBED_MODEL_NAME)
        embed_model = _EMBED_MODEL_CACHE
    docs = load_all_docs()

    all_results: List[Dict] = []
    strategy_chunk_stats: Dict[str, Dict] = {}

    for strategy in strategies:
        print(f"\n{'='*60}")
        print(f"Strategy {strategy['id']}: {strategy['name']}")
        print(f"{'='*60}")

        print("Building chunks...")
        chunks = build_chunks(strategy, docs)
        tok_counts = [c["n_tokens"] for c in chunks]
        print(f"  {len(chunks)} chunks | avg={np.mean(tok_counts):.1f} | min={min(tok_counts)} | max={max(tok_counts)}")

        strategy_chunk_stats[strategy["id"]] = {
            "n_chunks": len(chunks),
            "avg_tokens": float(np.mean(tok_counts)),
            "min_tokens": int(min(tok_counts)),
            "max_tokens": int(max(tok_counts)),
            "std_tokens": float(np.std(tok_counts)),
        }

        print("Embedding chunks...")
        chunk_embeddings = embed_chunks(chunks, embed_model)

        def _eval_question(q):
            print(f"  [{q['id']}] {q['question'][:65]}...")
            retrieved = retrieve_top_k(q["question"], chunks, chunk_embeddings, embed_model, k=TOP_K)
            context_text = "\n\n".join(c["text"] for c in retrieved)

            # Run answer + faithfulness + relevance in parallel
            with ThreadPoolExecutor(max_workers=3) as ex:
                f_answer = ex.submit(generate_answer, api_key, q["question"], retrieved)
            answer, tok_in, tok_out = f_answer.result()

            with ThreadPoolExecutor(max_workers=2) as ex:
                f_faith = ex.submit(faithfulness_score, api_key, context_text, answer)
                f_rel   = ex.submit(relevance_score,   api_key, q["question"], answer)
                faith = f_faith.result()
                rel   = f_rel.result()

            rec5    = recall_at_k(chunks, retrieved, q, k=TOP_K)
            mrr_val = mrr(retrieved, q)
            f1      = token_f1(answer, q.get("expected", ""))
            cost    = compute_cost(tok_in, tok_out)

            print(f"    R@5={rec5:.2f} MRR={mrr_val:.2f} F1={f1:.2f} Faith={faith:.2f} Rel={rel:.2f} Cost=${cost:.4f}")
            return {
                "strategy_id": strategy["id"],
                "strategy_name": strategy["name"],
                "q_id": q["id"],
                "question": q["question"],
                "type": q["type"],
                "difficulty": q["difficulty"],
                "source_doc": q["source_doc"],
                "answer": answer,
                "expected": q.get("expected", ""),
                "retrieved_chunks": [{"id": c["chunk_id"], "text": c["text"]} for c in retrieved],
                "recall_at_5": rec5,
                "mrr": mrr_val,
                "f1": f1,
                "faithfulness": faith,
                "relevance": rel,
                "tokens_input": tok_in,
                "tokens_output": tok_out,
                "cost_usd": cost,
            }

        # Evaluate all questions for this strategy in parallel
        with ThreadPoolExecutor(max_workers=min(4, len(questions))) as ex:
            futures = {ex.submit(_eval_question, q): q for q in questions}
            for future in as_completed(futures):
                all_results.append(future.result())

    # ── Save raw results ──────────────────────────────────────────────────────
    raw_path = RESULTS_DIR / "raw_results.json"
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)
    print(f"\nRaw results → {raw_path}")

    # ── Aggregate per strategy ────────────────────────────────────────────────
    summary = []
    costs = []
    for strategy in strategies:
        sid = strategy["id"]
        rows = [r for r in all_results if r["strategy_id"] == sid]
        if not rows:
            continue
        stats = strategy_chunk_stats[sid]
        entry = {
            "strategy_id": sid,
            "strategy_name": strategy["name"],
            **stats,
            "recall_at_5": float(np.mean([r["recall_at_5"] for r in rows])),
            "mrr": float(np.mean([r["mrr"] for r in rows])),
            "f1": float(np.mean([r["f1"] for r in rows])),
            "faithfulness": float(np.mean([r["faithfulness"] for r in rows])),
            "relevance": float(np.mean([r["relevance"] for r in rows])),
            "cost_per_query": float(np.mean([r["cost_usd"] for r in rows])),
        }
        summary.append(entry)
        costs.append(entry["cost_per_query"])

    max_cost = max(costs) if costs else 1.0
    for s in summary:
        norm_cost = s["cost_per_query"] / max_cost if max_cost > 0 else 0.0
        s["composite_score"] = (
            s["recall_at_5"] * 0.20
            + s["mrr"] * 0.15
            + s["f1"] * 0.15
            + s["faithfulness"] * 0.15
            + s["relevance"] * 0.10
            + (1 - norm_cost) * 0.10
        )

    summary_path = RESULTS_DIR / "summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    print(f"Summary     → {summary_path}")

    # ── Export Excel ──────────────────────────────────────────────────────────
    xlsx_path = RESULTS_DIR / "benchmark_results.xlsx"
    export_excel(all_results, summary, xlsx_path)
    print(f"Excel       → {xlsx_path}")

    # ── Signal completion to UI ───────────────────────────────────────────────
    (RESULTS_DIR / "_benchmark_done.txt").write_text("done")

    # ── Print final table ─────────────────────────────────────────────────────
    best_sid = max(summary, key=lambda x: x["composite_score"])["strategy_id"]
    print("\n" + "="*95)
    print(f"{'Strat':<6} {'N_chunks':>8} {'AvgTok':>7} {'R@5':>6} {'MRR':>6} {'F1':>6} {'Faith':>7} {'Rel':>6} {'Cost':>8} {'Composite':>10}")
    print("-"*95)
    for s in summary:  # already filtered to selected strategies
        marker = "  ◀ WINNER" if s["strategy_id"] == best_sid else ""
        print(
            f"{s['strategy_id']:<6} {s['n_chunks']:>8} {s['avg_tokens']:>7.1f} "
            f"{s['recall_at_5']:>6.3f} {s['mrr']:>6.3f} {s['f1']:>6.3f} "
            f"{s['faithfulness']:>7.3f} {s['relevance']:>6.3f} "
            f"{s['cost_per_query']:>8.4f} {s['composite_score']:>10.4f}{marker}"
        )


def export_excel(all_results: List[Dict], summary: List[Dict], path: Path):
    wb = openpyxl.Workbook()
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BOLD = Font(bold=True)

    best_sid = max(summary, key=lambda x: x["composite_score"])["strategy_id"]

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "Summary"
    h1 = ["Strategy", "N_chunks", "Avg_tokens", "Recall@5", "MRR", "F1",
          "Faithfulness", "Relevance", "Cost/query", "Composite_score"]
    ws1.append(h1)
    for c in ws1[1]:
        c.font = BOLD
    for s in summary:
        row = [
            s["strategy_id"], s["n_chunks"], round(s["avg_tokens"], 1),
            round(s["recall_at_5"], 4), round(s["mrr"], 4), round(s["f1"], 4),
            round(s["faithfulness"], 4), round(s["relevance"], 4),
            round(s["cost_per_query"], 6), round(s["composite_score"], 4),
        ]
        ws1.append(row)
        if s["strategy_id"] == best_sid:
            for cell in ws1[ws1.max_row]:
                cell.fill = GREEN

    # Sheet 2 — Per Question
    ws2 = wb.create_sheet("Per Question")
    h2 = ["Strategy", "Q_ID", "Question", "Type", "Difficulty",
          "Recall@5", "MRR", "F1", "Faithfulness", "Relevance", "Cost"]
    ws2.append(h2)
    for c in ws2[1]:
        c.font = BOLD
    for r in all_results:
        ws2.append([
            r["strategy_id"], r["q_id"], r["question"], r["type"], r["difficulty"],
            round(r["recall_at_5"], 4), round(r["mrr"], 4), round(r["f1"], 4),
            round(r["faithfulness"], 4), round(r["relevance"], 4), round(r["cost_usd"], 6),
        ])

    # Sheet 3 — Chunk Stats
    ws3 = wb.create_sheet("Chunk Stats")
    h3 = ["Strategy", "n_chunks", "avg_tokens", "min_tokens", "max_tokens", "std_tokens"]
    ws3.append(h3)
    for c in ws3[1]:
        c.font = BOLD
    for s in summary:
        ws3.append([
            s["strategy_id"], s["n_chunks"], round(s["avg_tokens"], 1),
            s["min_tokens"], s["max_tokens"], round(s["std_tokens"], 1),
        ])

    wb.save(str(path))


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--strategies", nargs="+", default=None,
        help="Strategy IDs to run, e.g. --strategies C E. Defaults to all.",
    )
    args = parser.parse_args()
    if args.strategies:
        valid_ids = {s["id"] for s in STRATEGIES}
        unknown = set(args.strategies) - valid_ids
        if unknown:
            raise ValueError(f"Unknown strategy IDs: {unknown}. Valid: {valid_ids}")
        active_strategies = [s for s in STRATEGIES if s["id"] in args.strategies]
        run_benchmark(strategies=active_strategies)
    else:
        run_benchmark()
